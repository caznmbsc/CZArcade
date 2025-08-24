from openpyxl import load_workbook
import json
import requests
from io import BytesIO
from bs4 import BeautifulSoup

print("Retrieving XLSX...")
sheetsLink = "https://docs.google.com/spreadsheets/d/1Hnb113j20olLsh-SR5zFUsQNV95BMmeXmQOkATtgleI/export?format=xlsx"
response = requests.get(sheetsLink)
attempts = 0
while response.status_code != 200:
    attempts += 1
    if attempts >= 10:
        print("Too many Fails. Terminating.")
        exit(1)
    print("Code: " + str(response.status_code) + ". Failed to Download Google Sheets File. Retrying...")
    response = requests.get(sheetsLink)
currentWorkbook = load_workbook(filename=BytesIO(response.content), data_only=True)
print("Downloaded Google Sheets File.\n") 

print("Making JSON...")
imageDictionary = {}
data = {}
#Make Combined Section for all Inventory
data["Inventory"] = []
for sheetName in currentWorkbook.sheetnames:
    if sheetName not in ["Rejects_Inventory", "Running_Inventory", "Systems", "Genres"]:
        continue
    print("\n-----------------------")
    print(f"Parsing: {sheetName}")
    print("-----------------------")
    sheet = currentWorkbook[sheetName]
    sheetData = []
    for row in sheet.iter_rows(min_row=1, values_only=False):
        rowData = {}
        for cell in row:
            if (cell.value == row[0].value) and (cell.value != None):
                print(cell.value)
            value = cell.value
            if cell.hyperlink:
                imageURL = "media/noImage.png"
                if cell.hyperlink.target in imageDictionary:
                    #If we've already run this, then get it from cache
                    imageURL = imageDictionary[cell.hyperlink.target]
                    print("\tLinked from Cache.")
                elif cell.hyperlink.target and ("launchbox" in cell.hyperlink.target):
                    #If there is a LaunchBox hyperlink, go get an image 
                    gameDBResponse = requests.get(cell.hyperlink.target)
                    attempts = 0
                    while gameDBResponse.status_code != 200:
                        attempts += 1
                        if attempts >= 10:
                            print("Too many Fails. Terminating.")
                            exit(1)
                        print("Code: " + str(gameDBResponse.status_code) + ". Failed to Download Google Sheets File. Retrying...")
                        gameDBResponse = requests.get(sheetsLink)
                    print("\tHTML recieved.")
                    dbPageHTML = BeautifulSoup(gameDBResponse.text, "html.parser")
                    if (dbPageHTML.find(string="Screenshot - Gameplay") and
                        dbPageHTML.find(string="Screenshot - Gameplay").parent.find_next_sibling("div") and
                        dbPageHTML.find(string="Screenshot - Gameplay").parent.find_next_sibling("div").find("div") and
                        dbPageHTML.find(string="Screenshot - Gameplay").parent.find_next_sibling("div").find("div").find("img")
                    ):
                        #Get gameplay image if it exists
                        imageHolder = dbPageHTML.find(string="Screenshot - Gameplay").parent.find_next_sibling("div")
                        imageURL = imageHolder.find("div").find("img")["src"]
                        print("\tFetched Image.")
                    elif (dbPageHTML.find("section", {"class": "game-details-content"}) and 
                        dbPageHTML.find("section", {"class": "game-details-content"}).find_all("article")[1] and
                        dbPageHTML.find("section", {"class": "game-details-content"}).find_all("article")[1].find("img")
                    ):
                        #If no gameplay image exists, get any image from the media section
                        imageURL = dbPageHTML.find("section", {"class": "game-details-content"}).find_all("article")[1].find("img")["src"]
                        print("\tFetched Image.")
                    imageDictionary[cell.hyperlink.target] = imageURL
                value = {"text": value, "hyperlink": cell.hyperlink.target, "imageURL": imageURL}
            rowData[cell.column_letter] = value
        sheetData.append(rowData)
    if sheetName in ["Rejects_Inventory", "Running_Inventory"]:
        #Concatenate Rejects and Running Inventories
        data["Inventory"] += sheetData
    else:
        data[sheetName] = sheetData

with open("CZARCADE.json", "w", encoding="utf-8") as jsonFile:
    json.dump(data, jsonFile, ensure_ascii=False, indent=2)
print("JSON Created.")